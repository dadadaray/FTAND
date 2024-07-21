import pandas as pd
from tqdm import tqdm
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# 读取Excel文件中的两个表
excel_file = 'datas.xlsx'
forest_food_data = pd.read_excel(excel_file, sheet_name='森林食物数据')
final_data = pd.read_excel(excel_file, sheet_name='final_data')

# 将‘final_data’中的第G列和‘森林食物数据’中的第A列转换为字符串类型，以确保部分匹配能够进行
final_data['G列'] = final_data.iloc[:, 6].astype(str)
forest_food_data['A列'] = forest_food_data.iloc[:, 0].astype(str)

# 创建一个字典来存储final_data中G列的集合及其对应的行数据
final_data_dict = {}
for _, f_row in final_data.iterrows():
    f_row_set = set(f_row['G列'])
    for char in f_row_set:
        if char not in final_data_dict:
            final_data_dict[char] = []
        final_data_dict[char].append((f_row['G列'], f_row.iloc[0:6].values))


# 定义一个函数来查找匹配的行
def match_and_fill(row, final_data_dict):
    row_value = row['A列']
    row_set = set(row_value)
    is_partial_match = False

    # 1. 完全匹配优先
    for char in row_set:
        if char in final_data_dict:
            for g_value, values in final_data_dict[char]:
                if row_value == g_value:
                    return list(values) + [g_value, "完全匹配"]

    # 2. 检查至少两个字符相等的匹配
    match_count = {}
    g_value_map = {}
    for char in row_set:
        if char in final_data_dict:
            for g_value, values in final_data_dict[char]:
                values_tuple = tuple(values)
                if values_tuple not in match_count:
                    match_count[values_tuple] = 0
                    g_value_map[values_tuple] = g_value
                match_count[values_tuple] += 1
                if match_count[values_tuple] >= 2:  # 至少有两个字相等
                    is_partial_match = True
                    return list(values) + [g_value, "不完全匹配"]

    # 3. 无匹配项
    if is_partial_match:
        return [None] * 6 + ["无匹配项", "无匹配项"]  # 如果没有匹配项，返回空值和“无匹配项”
    else:
        return [None] * 6 + ["无匹配项", "无匹配项", "不完全匹配"]


# 初始化一个空列表来存储匹配的结果
matches = []

# 使用tqdm进度条
for _, row in tqdm(forest_food_data.iterrows(), total=forest_food_data.shape[0]):
    matches.append(match_and_fill(row, final_data_dict))

# 将匹配的值填入‘森林食物数据’的第B-I列
forest_food_data.iloc[:, 1:9] = pd.DataFrame(matches, index=forest_food_data.index)

# 保存结果到新的Excel文件
temp_file = 'temp_forest_food_data.xlsx'
forest_food_data.to_excel(temp_file, sheet_name='森林食物数据', index=False)

print("匹配结果已保存到临时文件 'temp_forest_food_data.xlsx'")

# 加载临时Excel文件
wb = load_workbook(temp_file)
ws = wb['森林食物数据']

# 定义红色填充样式
red_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")

# 遍历每一行，标记不完全匹配的单元格
for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    is_partial_match = False
    for cell in row[1:8]:  # 第2列到第8列
        if cell.value == "不完全匹配":
            is_partial_match = True
            break
    if is_partial_match:
        for cell in row:
            cell.fill = red_fill

# 保存最终结果到新的Excel文件
final_file = 'updated_forest_food_data.xlsx'
wb.save(final_file)

print("操作完成，文件已保存为 'updated_forest_food_data.xlsx'")
